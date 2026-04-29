# -*- coding: utf-8 -*-
"""
日本の1兆円市場マップ (FY2024)
- 親産業 → サブセグメント の階層で、1兆円規模の国内市場を横断列挙
- 各サブセグメントについて、上位5社の「当該セグメント売上」を併記
- セグメント開示が無い企業は載せない (空欄で可)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== Styles (workspace/fermi_billmaintenance.py の定数を流用) =====
BLUE   = Font(color='0070C0', name='Yu Gothic', size=10)
BLACK  = Font(color='000000', name='Yu Gothic', size=10)
BLACK_B= Font(color='000000', name='Yu Gothic', size=10, bold=True)
HEADER = Font(color='FFFFFF', name='Yu Gothic', size=11, bold=True)
TITLE  = Font(color='000000', name='Yu Gothic', size=14, bold=True)

HEADER_FILL = PatternFill('solid', fgColor='305496')
PARENT_FILL = PatternFill('solid', fgColor='D9E1F2')
ALT_FILL    = PatternFill('solid', fgColor='F2F7FB')

thin = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FMT_OKU = '#,##0'
FMT_TRN = '0.0"兆円"'

# ===== Data (各リサーチagentの出力をここに統合) =====
# Schema:
#   {parent, segment, size_trillion, size_year, size_source, size_url, size_note,
#    companies: [{name, revenue_oku, fy, source, url, note?}, ...],
#    note}
MARKETS = [
    # ===== 建設 =====
    {
        'parent': '建設',
        'segment': '総合建設 (大手ゼネコン)',
        'size_trillion': 19.0,
        'size_year': 'FY2024',
        'size_source': '国交省 建設工事受注動態統計 大手50社受注高 18兆9558億円 (令和6年度)',
        'size_url': 'https://www.mlit.go.jp/report/press/joho04_hh_001296.html',
        'companies': [
            {'name': '鹿島建設', 'revenue_oku': 29118, 'fy': 'FY2024', 'source': '2025年3月期決算短信 連結売上高',
             'url': 'https://www.kajima.co.jp/ir/finance/pdf/kessan-20250514-j.pdf'},
            {'name': '大林組', 'revenue_oku': 26201, 'fy': 'FY2024', 'source': '2025年3月期決算 連結売上高',
             'url': 'https://ir.obayashi.co.jp/ja/ir/data/results.html'},
            {'name': '大成建設', 'revenue_oku': 21542, 'fy': 'FY2024', 'source': '2025年3月期決算短信 連結売上高',
             'url': 'https://www.taisei.co.jp/ir/data/tanshin/'},
            {'name': '清水建設', 'revenue_oku': 19443, 'fy': 'FY2024', 'source': '2025年3月期決算短信 連結売上高',
             'url': 'https://www.shimz.co.jp/ir/financial/index.html'},
            {'name': '竹中工務店', 'revenue_oku': 16001, 'fy': 'FY2024', 'source': '2024年12月期決算 連結売上高 (非上場、有報公表)',
             'url': 'https://www.takenaka.co.jp/news/2025/02/06/'},
        ],
        'note': '5社合計連結売上 約11.2兆円。市場規模は大手50社受注高ベース。各社売上は連結売上 (海外・不動産等含む)。',
    },
    {
        'parent': '建設',
        'segment': '空調・衛生設備工事',
        'size_trillion': 1.9,
        'size_year': 'FY2024',
        'size_source': '国交省 設備工事業受注高調査 主要20社の管工事受注高 1兆8544億円 (令和6年度)',
        'size_url': 'https://www.mlit.go.jp/sogoseisaku/jouhouka/sosei_jouhouka_tk4_000010.html',
        'companies': [
            {'name': '高砂熱学工業', 'revenue_oku': 3817, 'fy': 'FY2024', 'source': '2025年3月期決算短信 連結売上高',
             'url': 'https://www.tte-net.com/ir/library/results/'},
            {'name': '新菱冷熱工業', 'revenue_oku': 3084, 'fy': 'FY2024 (2024/9期)', 'source': '2024年9月期決算 (非上場、財務情報公表)',
             'url': 'https://www.shinryo.com/news/20241209.html'},
            {'name': 'ダイダン', 'revenue_oku': 2627, 'fy': 'FY2024', 'source': '2025年3月期決算短信 完成工事高',
             'url': 'https://www.daidan.co.jp/wp/wp-content/uploads/2025/05/3add1894ebea6434b56c7b77d15c3e98.pdf'},
            {'name': '三機工業', 'revenue_oku': 2531, 'fy': 'FY2024', 'source': '2025年3月期決算短信 連結売上高 (過去最高)',
             'url': 'https://www.sanki.co.jp/ir/'},
            {'name': '新日本空調', 'revenue_oku': 1376, 'fy': 'FY2024', 'source': '2025年3月期決算 連結売上高',
             'url': 'https://www.snk.co.jp/ir/library/'},
        ],
        'note': '市場規模は大手20社受注高ベース。業界全体ベースだと約3兆円規模。朝日工業社 (919億円) は次点。',
    },
    {
        'parent': '建設',
        'segment': '住宅 (戸建+マンション)',
        'size_trillion': 16.5,
        'size_year': '2024年(暦年)',
        'size_source': '国交省 建築着工統計 (新設住宅着工 79.2万戸、住宅工事費予定額ベース約16.5兆円)',
        'size_url': 'https://www.mlit.go.jp/sogoseisaku/jouhouka/sosei_jouhouka_tk4_000002.html',
        'companies': [
            {'name': '大和ハウス工業', 'revenue_oku': 27899, 'fy': 'FY2024', 'source': '2025年3月期 戸建+賃貸+マンション3セグ合算',
             'url': 'https://www.daiwahouse.co.jp/about/release/house/pdf/release_20250513.pdf'},
            {'name': '飯田グループHD', 'revenue_oku': 12091, 'fy': 'FY2024', 'source': '2025年3月期 戸建分譲事業セグメント (38,627戸)',
             'url': 'https://www.ighd.co.jp/ir/library/document.html'},
            {'name': 'オープンハウスG', 'revenue_oku': 5902, 'fy': 'FY2024 (2024/9期)', 'source': '2024年9月期 戸建関連事業',
             'url': 'https://openhouse-group.co.jp/ir/upload_file/m005-m005_07/244q_kessan.pdf'},
            {'name': '住友林業', 'revenue_oku': 5423, 'fy': 'FY2024 (2024/12期)', 'source': '2024年12月期 国内住宅事業',
             'url': 'https://sfc.jp/information/ir/library/statements/pdf/2024-4q_ren.pdf'},
            {'name': '積水ハウス', 'revenue_oku': 4790, 'fy': 'FY2024 (2025/1期)', 'source': '2025年1月期 戸建住宅セグメント',
             'url': 'https://www.sekisuihouse.co.jp/company/financial/library/ir_document/2024/2024_kessan/t20250306.pdf'},
        ],
        'note': '大和ハウスは戸建/賃貸/マンション3セグ合算。積水ハウスの請負型ビジネス全体は1.35兆円。ミサワホームは戸建単独額不開示で除外。',
    },
    {
        'parent': '建設',
        'segment': 'リフォーム・リノベーション',
        'size_trillion': 7.35,
        'size_year': '2024年(暦年)',
        'size_source': '矢野経済研究所 住宅リフォーム市場調査2025 — 7兆3470億円 (前年比0.5%減)',
        'size_url': 'https://www.yano.co.jp/press-release/show/press_id/3877',
        'companies': [
            {'name': '積水ハウスG', 'revenue_oku': 1750, 'fy': 'FY2024', 'source': 'リフォーム産業新聞 売上高ランキング2024 1位',
             'url': 'https://www.reform-online.jp/news/reform-shop/25496.php'},
            {'name': '大和ハウスG', 'revenue_oku': 1730, 'fy': 'FY2024', 'source': 'リフォーム産業新聞 売上高ランキング2024 2位',
             'url': 'https://www.reform-online.jp/news/reform-shop/25496.php'},
            {'name': '住友不動産G', 'revenue_oku': 1131, 'fy': 'FY2024', 'source': 'リフォーム産業新聞 売上高ランキング2024 3位',
             'url': 'https://www.reform-online.jp/news/reform-shop/25496.php'},
            {'name': '積水化学G', 'revenue_oku': 1033, 'fy': 'FY2024', 'source': 'リフォーム産業新聞 売上高ランキング2024 4位',
             'url': 'https://www.reform-online.jp/news/reform-shop/25496.php'},
            {'name': 'ヤマダHD', 'revenue_oku': 886, 'fy': 'FY2024', 'source': 'リフォーム産業新聞 売上高ランキング2024 5位',
             'url': 'https://www.reform-online.jp/news/reform-shop/25496.php'},
        ],
        'note': '※リフォームは有報セグメント開示が稀のため、業界紙 (リフォーム産業新聞) の調査ベース売上を採用。',
    },

    # ===== 自動車 =====
    {
        'parent': '自動車',
        'segment': '自動車内装・シート部品',
        'size_trillion': 3.5,
        'size_year': 'FY2023',
        'size_source': 'JAPIA 自動車部品出荷動向調査2023年度 (出荷総額21.07兆円のうち車体・内装系約16-17%)',
        'size_url': 'https://www.japia.or.jp/files/user/japia/research/shukka/2023shukka.pdf',
        'companies': [
            {'name': 'トヨタ紡織', 'revenue_oku': 19542, 'fy': 'FY2024', 'source': '2025年3月期 連結売上収益 (シート・内装一貫)',
             'url': 'https://www.toyota-boshoku.com/jp/company/_assets/upload/2025_brief.pdf'},
            {'name': 'テイ・エス テック', 'revenue_oku': 4605, 'fy': 'FY2024', 'source': '2025年3月期 売上収益 (四輪事業=シート 93%)',
             'url': 'https://www.tstech.co.jp/ir/highlight/'},
            {'name': '豊田合成', 'revenue_oku': 3465, 'fy': 'FY2024', 'source': '2025年3月期 内外装事業売上 (連結×33%)',
             'url': 'https://www.toyoda-gosei.co.jp/ir/document/accounts_data/'},
            {'name': 'タチエス', 'revenue_oku': 2853, 'fy': 'FY2024', 'source': '2025年3月期 連結売上高 (シート専業)',
             'url': 'https://www.tachi-s.co.jp/ir/library/brief_note.html'},
        ],
        'note': '市場規模はJAPIA FY2023ベース (FY2024品目別は2026年3月公表予定)。林テレンプは非上場で除外。各社売上は連結 (グローバル) ベース。',
    },
    {
        'parent': '自動車',
        'segment': '自動車駆動・電動化部品',
        'size_trillion': 6.5,
        'size_year': 'FY2023',
        'size_source': 'JAPIA 自動車部品出荷動向調査2023年度 (出荷総額21.07兆円のうち駆動・伝導+電気装置 約30%)',
        'size_url': 'https://www.japia.or.jp/files/user/japia/research/shukka/2023shukka.pdf',
        'companies': [
            {'name': 'アイシン', 'revenue_oku': 26801, 'fy': 'FY2024', 'source': '2025年3月期 パワートレイン関連製品 売上 (構成比54.7%)',
             'url': 'https://www.aisin.com/jp/news/2025/009016.html'},
            {'name': 'ニデック', 'revenue_oku': 6800, 'fy': 'FY2024', 'source': '2025年3月期 車載事業セグメント (構成比約26%)',
             'url': 'https://www.nidec.com/files/user/www-nidec-com/ir/library/earnings/2025/FY24Q4_3_jp.pdf'},
            {'name': '武蔵精密工業', 'revenue_oku': 3472, 'fy': 'FY2024', 'source': '2025年3月期 連結売上高 (差動装置・ギア専業)',
             'url': 'https://www.musashi.co.jp/ir/'},
            {'name': 'エクセディ', 'revenue_oku': 3000, 'fy': 'FY2024', 'source': '2025年3月期 連結売上収益 (MT/AT/トルコン)',
             'url': 'https://exf.exedy.com/ja/assets/pdf/stockholder/FY2024_FinancialStatement.pdf'},
        ],
        'note': 'デンソー はパワートレイン単独セグメント開示無 (2024年4月組織変更後) で除外。ジヤトコは非上場で除外。各社売上はグローバル連結。',
    },
    {
        'parent': '自動車',
        'segment': 'タイヤ',
        'size_trillion': 1.4,
        'size_year': '2024年(暦年)',
        'size_source': 'JATMA 2024年自動車タイヤ国内需要実績 (1.03億本、新車3,648万本+市販6,350万本)',
        'size_url': 'https://www.jatma.or.jp/tyre_industry/production_sales.html',
        'companies': [
            {'name': 'ブリヂストン', 'revenue_oku': 12261, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 日本セグメント売上 (多角化事業含む)',
             'url': 'https://www.bridgestone.co.jp/ir/library/result/pdf/r6_4_4.pdf'},
            {'name': '住友ゴム工業', 'revenue_oku': 10464, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 タイヤ事業セグメント (グローバル)',
             'url': 'https://www.srigroup.co.jp/ir/library/financial-report/dvql4p000000q694-att/2024_end_tanshin.pdf'},
            {'name': '横浜ゴム', 'revenue_oku': 9809, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 タイヤ事業セグメント (グローバル)',
             'url': 'https://www.y-yokohama.com/release/?id=4490&lang=ja'},
            {'name': 'TOYO TIRE', 'revenue_oku': 5198, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 タイヤ事業セグメント (過去最高)',
             'url': 'https://www.toyotires.co.jp/ir/'},
        ],
        'note': 'JATMA市場規模は本数ベースのみ公表 (金額は本数×平均単価で試算、約1.4兆円)。各社売上はグローバル合算 (国内のみ抽出不可)。',
    },

    # ===== 小売 =====
    {
        'parent': '小売',
        'segment': '食品スーパー',
        'size_trillion': 12.9,
        'size_year': '2024年(暦年)',
        'size_source': '経産省 商業動態統計 2024年確報 (スーパー販売額のうち飲食料品 約12兆8968億円)',
        'size_url': 'https://www.meti.go.jp/statistics/tyo/syoudou/result/kakuho_2.html',
        'companies': [
            {'name': 'ライフコーポレーション', 'revenue_oku': 8505, 'fy': 'FY2024 (2025/2)', 'source': '2025年2月期決算短信 連結営業収益',
             'url': 'https://www.lifecorp.jp/vc-files/pdf/ir/financial_results/2025.4Q_sankou.pdf'},
            {'name': 'USMH', 'revenue_oku': 8113, 'fy': 'FY2024 (2025/2)', 'source': '2025年2月期決算 連結営業収益 (マルエツ/カスミ/MV/いなげや)',
             'url': 'https://www.usmh.co.jp/wp-content/themes/usmh/files/20250410/Ust_20250410_1540.pdf'},
            {'name': 'ヤオコー', 'revenue_oku': 7364, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結営業収益',
             'url': 'https://www.yaoko-net.com/ir/'},
            {'name': 'アークス', 'revenue_oku': 6083, 'fy': 'FY2024 (2025/2)', 'source': '2025年2月期決算 連結営業収益 (北海道・東北のSM)',
             'url': 'https://www.arcs-g.co.jp/vc-files/arcs-g/ir/pdf/2024/arcs_202502full-yearfinancialresults_2025041404.pdf'},
            {'name': 'バローHD', 'revenue_oku': 4912, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 SM事業セグメント (Q3累計から外挿)',
             'url': 'https://valorholdings.co.jp/wp-content/uploads/published/wp_re_68_tanshin_20250513.pdf'},
        ],
        'note': '非上場ベイシア(7000億規模)、トライアルHDは別カウント。バローはSMセグのみ抽出 (HC等を除く)。',
    },
    {
        'parent': '小売',
        'segment': 'コンビニエンスストア',
        'size_trillion': 11.8,
        'size_year': '2024年(暦年)',
        'size_source': 'JFA コンビニエンスストア統計2024年集計 全店売上 11兆7953億円',
        'size_url': 'https://www.jfa-fc.or.jp/particle/320.html',
        'companies': [
            {'name': 'セブン-イレブン・ジャパン', 'revenue_oku': 53698, 'fy': 'FY2024 (2025/2)', 'source': 'セブン&アイHD 国内CVS事業 SEJ単体チェーン全店売上',
             'url': 'https://www.7andi.com/ir/library/co_financial/2025/convenience_store'},
            {'name': 'ファミリーマート', 'revenue_oku': 32439, 'fy': 'FY2024 (2025/2)', 'source': 'ファミリーマート 単体チェーン全店売上 (伊藤忠100%子会社)',
             'url': 'https://www.family.co.jp/content/dam/family/company/familymart/overview/250409_gyosekigaikyo.pdf'},
            {'name': 'ローソン', 'revenue_oku': 28919, 'fy': 'FY2024 (2025/2)', 'source': '2025年2月期決算 国内CVSチェーン全店売上',
             'url': 'https://www.lawson.co.jp/company/ir/library/materials/'},
            {'name': 'セイコーマート', 'revenue_oku': 2300, 'fy': 'FY2024', 'source': 'セコマ (非上場、業界紙参考値)',
             'url': 'https://www.seicomart.co.jp/'},
            {'name': 'ミニストップ', 'revenue_oku': 1934, 'fy': 'FY2024 (2025/2)', 'source': 'イオン傘下 2025年2月期通期売上 (推定)',
             'url': 'https://www.ministop.co.jp/corporate/ir/'},
        ],
        'note': 'チェーン全店売上ベース (JFA市場規模との整合性優先)。セブン&アイ国内CVS事業のIFRS収益は0.9兆円 (定義異なる)。',
    },
    {
        'parent': '小売',
        'segment': 'ドラッグストア',
        'size_trillion': 10.0,
        'size_year': 'FY2024',
        'size_source': '日本チェーンドラッグストア協会(JACDS) 第25回実態調査 — 10兆307億円 (初の10兆円台)',
        'size_url': 'https://jacds.gr.jp/',
        'companies': [
            {'name': 'ウエルシアHD', 'revenue_oku': 12850, 'fy': 'FY2024 (2025/2)', 'source': '2025年2月期決算短信 連結売上高',
             'url': 'https://data.swcms.net/file/welcia/dam/jcr:a2d912e4-a80d-4d28-ae2d-c1d2d9bacf82/140120250408510937.pdf'},
            {'name': 'マツキヨココカラ&カンパニー', 'revenue_oku': 10616, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上高',
             'url': 'https://www.matsukiyococokara.com/ir/library/results/'},
            {'name': 'コスモス薬品', 'revenue_oku': 10114, 'fy': 'FY2024 (2025/5)', 'source': '2025年5月期決算 (初の1兆円超)',
             'url': 'https://www.cosmospc.co.jp/ir/ir_financial/pdf/4ecabde593934f7d1d520dcce64ec5c906c71e1f.pdf'},
            {'name': 'ツルハHD', 'revenue_oku': 8456, 'fy': 'FY2024 (2025/2、9.5ヶ月)', 'source': '2025年2月期 (決算期変更9.5ヶ月、年換算約1.07兆)',
             'url': 'https://www.tsuruha-hd.com/ir/library/results.html'},
            {'name': 'サンドラッグ', 'revenue_oku': 8018, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上高',
             'url': 'https://www.sundrug.co.jp/ir/irdata/results'},
        ],
        'note': 'スギHD (8780億) は6位相当。ウエルシア・ツルハは2027年9月にイオン主導で経営統合予定。',
    },
    {
        'parent': '小売',
        'segment': '家電量販',
        'size_trillion': 10.3,
        'size_year': '2024年(暦年)',
        'size_source': '経産省 商業動態統計 2024年家電大型専門店販売額 10兆2920億円 (3年ぶり10兆円台)',
        'size_url': 'https://www.meti.go.jp/statistics/tyo/syoudou/result/kakuho_2.html',
        'companies': [
            {'name': 'ヤマダHD', 'revenue_oku': 13089, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 デンキセグメント売上 (住建・金融除く)',
             'url': 'https://www.yamada-holdings.jp/ir/kessan/2025/250508.pdf'},
            {'name': 'ビックカメラ', 'revenue_oku': 9745, 'fy': 'FY2024 (2025/8)', 'source': '2025年8月期決算 連結売上高 (コジマ・ソフマップ含む、過去最高)',
             'url': 'https://www.biccamera.co.jp/ir/financial/results/'},
            {'name': 'ヨドバシカメラ', 'revenue_oku': 8162, 'fy': 'FY2024 (2025/3)', 'source': 'ヨドバシ (非上場、業界紙推計値)',
             'url': 'https://diamond-rm.net/market/accounting/494340/'},
            {'name': 'エディオン', 'revenue_oku': 7681, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上高',
             'url': 'https://www.edion.co.jp/system/files/ir-library/pdf/ja/2025-05/20250509_ir_1.pdf'},
            {'name': 'ケーズHD', 'revenue_oku': 7380, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上高',
             'url': 'https://www.ksdenki.co.jp/ir/library/financial_summary/'},
        ],
        'note': 'ヤマダHDはデンキセグメントで他社と比較可能性確保。ヨドバシは非上場のため業界誌推計値。',
    },
    {
        'parent': '小売',
        'segment': 'EC (BtoC物販)',
        'size_trillion': 15.2,
        'size_year': '2024年(暦年)',
        'size_source': '経産省 令和6年度電子商取引市場調査 BtoC物販系15兆2194億円 (EC化率9.78%)',
        'size_url': 'https://www.meti.go.jp/press/2025/08/20250826005/20250826005.html',
        'companies': [
            {'name': '楽天グループ', 'revenue_oku': 59550, 'fy': 'FY2024 (2024/12)', 'source': '楽天G 2024年度 国内EC流通総額 (楽天市場主体)',
             'url': 'https://corp.rakuten.co.jp/news/press/2025/0214_01.html'},
            {'name': 'LINEヤフー', 'revenue_oku': 43766, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 国内eコマース取扱高 (Yショッピング/ZOZO/LOHACO等)',
             'url': 'https://www.lycorp.co.jp/ja/ir/news/auto_20250502530042/pdfFile.pdf'},
            {'name': 'アマゾンジャパン', 'revenue_oku': 38000, 'fy': 'FY2024 (2024/12、推計)', 'source': 'Amazon.com 10-K 国別売上 Japan ($25.4B、150円換算)',
             'url': 'https://ir.aboutamazon.com/sec-filings/default.aspx'},
            {'name': 'ZOZO', 'revenue_oku': 6144, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 商品取扱高',
             'url': 'https://corp.zozo.com/ir/files/pdf/74ae228c4454adeea73d42852e9975a3acb81106.pdf'},
            {'name': 'メルカリ', 'revenue_oku': 1926, 'fy': 'FY2024 (2025/6)', 'source': '2025年6月期 連結売上収益 (国内GMV約1.1兆円)',
             'url': 'https://about.mercari.com/ir/'},
        ],
        'note': '各社「流通総額」「取扱高」「売上」「GMV」の定義が混在。楽天・LINEヤフー・Amazonの3強で物販EC市場の大半を寡占。',
    },

    # ===== 食品・飲料 =====
    {
        'parent': '食品・飲料',
        'segment': '飲料 (清涼飲料+ビール+乳飲料)',
        'size_trillion': 6.0,
        'size_year': '2024年(暦年)',
        'size_source': '全国清涼飲料連合会(4.73兆) + ビール酒造組合(1.06兆) + 牛乳乳製品統計(0.2兆) 合算約6兆円',
        'size_url': 'https://www.j-sda.or.jp/statistically-information/pdf/2024jsda_databook.pdf',
        'companies': [
            {'name': 'コカ・コーラBJH', 'revenue_oku': 8824, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期決算短信 連結売上収益',
             'url': 'https://www.ccbj-holdings.com/pdf/irinfo/237_1.pdf'},
            {'name': 'サントリー食品インターナショナル', 'revenue_oku': 7352, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期決算 日本事業セグメント',
             'url': 'https://www.suntory.co.jp/softdrink/ir/library_earnings/upload/2024_4q_tanshin.pdf'},
            {'name': 'キリンビール (キリンHD)', 'revenue_oku': 6627, 'fy': 'FY2024 (2024/12)', 'source': 'キリンG 2024年度 キリンビール売上収益',
             'url': 'https://pdf.irpocket.com/C2503/A7NF/uhQ5/gt8R.pdf'},
            {'name': 'サッポロHD (酒類+食品飲料)', 'revenue_oku': 5061, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 セグメント別売上 (酒類3882+食品飲料1179)',
             'url': 'https://www.sapporoholdings.jp/ir/library/factbook/items/2025__05_factbook.pdf'},
            {'name': '伊藤園', 'revenue_oku': 4538, 'fy': 'FY2024 (2024/4期)', 'source': '2024年4月期決算 連結売上高',
             'url': 'https://www.itoen.co.jp/wp-content/uploads/2023/12/20240603__shiryo.pdf'},
        ],
        'note': 'アサヒGHDは「日本」セグメント (酒類+飲料+食品一括) で純粋飲料の切出し不可のため除外。',
    },
    {
        'parent': '食品・飲料',
        'segment': '加工食品 (製造業出荷ベース)',
        'size_trillion': 31.1,
        'size_year': 'FY2023',
        'size_source': '矢野経済研究所 国内加工食品市場調査 (メーカー出荷ベース31.1兆円)。狭義 (冷凍1.3+即席麺0.8) では2-3兆円規模。',
        'size_url': 'https://www.yano.co.jp/press-release/show/press_id/3625',
        'companies': [
            {'name': '明治HD', 'revenue_oku': 9001, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 食品セグメント (乳製品・菓子・栄養含む広義)',
             'url': 'https://www.meiji.com/pdf/investor/library/settlement_2024_r04.pdf'},
            {'name': '日本ハム', 'revenue_oku': 4312, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 加工事業本部',
             'url': 'https://www.nipponham.co.jp/ir/library/briefing_session/2024.html'},
            {'name': 'ニチレイ', 'revenue_oku': 2909, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 加工食品事業セグメント',
             'url': 'https://www.nichirei.co.jp/sites/default/files/inline-images/ir/pdf_file/pres/240514kettsan_1.pdf'},
            {'name': 'マルハニチロ', 'revenue_oku': 1049, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 加工食品事業セグメント',
             'url': 'https://www.maruha-nichiro.co.jp/corporate/ir/library/pdf/20240516_4Q_presentation.pdf'},
            {'name': '東洋水産', 'revenue_oku': 1001, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 国内即席麺セグメント',
             'url': 'https://finance-frontend-pc-dist.west.edge.storage-yahoo.jp/disclosure/20240510/20240510588067.pdf'},
        ],
        'note': '味の素は調味料・食品セグが国内+海外合算のため除外。市場規模は加工食品全体出荷額 (31兆円)、狭義 (冷凍+即席) では2-3兆円規模。',
    },

    # ===== 外食 =====
    {
        'parent': '外食',
        'segment': '外食 (チェーン+全業態)',
        'size_trillion': 30.0,
        'size_year': '2024年(暦年)',
        'size_source': '日本フードサービス協会 2024年外食市場動向調査 (前年比108.4%)、外食産業市場規模約30兆円',
        'size_url': 'https://www.jfnet.or.jp/wp/wp-content/uploads/2025/09/nenkandata-2024pdf.pdf',
        'companies': [
            {'name': 'ゼンショーHD', 'revenue_oku': 9658, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 連結売上 (すき家・はま寿司等、国内外含む)',
             'url': 'https://www.zensho.co.jp/jp/ir/resource/pdf/24.5.14.zensho.all.pdf'},
            {'name': '日本マクドナルドHD', 'revenue_oku': 4055, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期決算 単一セグメント',
             'url': 'https://www.mcd-holdings.co.jp/ir/'},
            {'name': 'すかいらーくHD', 'revenue_oku': 4011, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期決算短信(IFRS)',
             'url': 'https://corp.skylark.co.jp/ir/library/brief_note/'},
            {'name': 'FOOD&LIFE COMPANIES', 'revenue_oku': 3611, 'fy': 'FY2024 (2024/9)', 'source': '2024年9月期 連結売上収益 (スシロー他)',
             'url': 'https://www.food-and-life.co.jp/wp-content/uploads/2024/12/medium_term_plan_2024.pdf'},
            {'name': 'コロワイド', 'revenue_oku': 2413, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 連結売上',
             'url': 'https://www.colowide.co.jp/ir/highlight/'},
        ],
        'note': 'ゼンショーは国内外合算。マクドナルドは単一セグメント。トリドール (丸亀製麺) は1148億円で次点。',
    },

    # ===== ヘルスケア =====
    {
        'parent': 'ヘルスケア',
        'segment': '医療用医薬品 (国内)',
        'size_trillion': 11.49,
        'size_year': 'FY2024',
        'size_source': 'IQVIA Japan Pharmaceutical Market 2024 / JPMA DATABOOK 2025 (薬価ベース11兆4874億円、3年連続最高)',
        'size_url': 'https://www.jpma.or.jp/news_room/issue/databook/ja/eo4se30000005nw2-att/DATABOOK2025.pdf',
        'companies': [
            {'name': '武田薬品工業 (日本)', 'revenue_oku': 5400, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 地域別売上 日本 (推定値含む)',
             'url': 'https://www.takeda.com/jp/investors/'},
            {'name': '第一三共 (JBU)', 'revenue_oku': 4770, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 ジャパンビジネスユニット',
             'url': 'https://www.daiichisankyo.co.jp/files/investors/library/quarterly_result/2024/4Q/FY2024Q4_Financial_Results_J.pdf'},
            {'name': '中外製薬', 'revenue_oku': 4611, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 国内製商品売上',
             'url': 'https://www.chugai-pharm.co.jp/news/detail/20250130170004_1464.html'},
            {'name': '大塚HD (医療関連-日本)', 'revenue_oku': 4200, 'fy': 'FY2024 (2024/12)', 'source': '2024年12月期 医療関連事業-地域別 (推定)',
             'url': 'https://www.otsuka.com/jp/ir/'},
            {'name': 'アステラス製薬 (日本)', 'revenue_oku': 1500, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 地域別売上 日本 (推定)',
             'url': 'https://www.astellas.com/jp/investors/financial-results-library'},
        ],
        'note': '武田・大塚・アステラスの日本セグメントFY2024は推定値含む (PDF直読不可)。塩野義・エーザイは2000億台で次点。',
    },
    {
        'parent': 'ヘルスケア',
        'segment': '介護サービス (民間)',
        'size_trillion': 11.94,
        'size_year': 'FY2024',
        'size_source': '厚労省 介護給付費等実態統計 (令和6年度) — 介護費用11兆9381億円 (前年度比+3.7%、過去最多)',
        'size_url': 'https://www.mhlw.go.jp/toukei/list/45-1.html',
        'companies': [
            {'name': 'ニチイHD', 'revenue_oku': 3080, 'fy': 'FY2022 (2023/3)', 'source': '電子公告 (2024年6月日本生命傘下で非上場化、最終公開値)',
             'url': 'https://www.nichiigakkan.co.jp/ir/notification.html'},
            {'name': 'ベネッセHD (介護・保育)', 'revenue_oku': 1600, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期決算補足 介護・保育セグメント (2024年MBOで上場廃止、最終値)',
             'url': 'https://www.benesse-hd.co.jp/ja/ir/doc/library/20240515_03j.pdf'},
            {'name': 'SOMPOケア', 'revenue_oku': 1500, 'fy': 'FY2023 (2024/3)', 'source': 'SOMPO HD 2024年3月期 介護・シニア事業セグメント',
             'url': 'https://www.sompo-hd.com/-/media/hd/files/doc/pdf/news2024/20240520_1.pdf?la=ja-JP'},
            {'name': 'ツクイHD', 'revenue_oku': 950, 'fy': 'FY2020 (2021/3)', 'source': '上場廃止前最終決算 (MBKパートナーズ傘下、現在非開示)',
             'url': 'https://www.tsukui.co.jp/'},
            {'name': 'ソラスト (介護)', 'revenue_oku': 580, 'fy': 'FY2023 (2024/3)', 'source': '2024年3月期 介護事業セグメント',
             'url': 'https://www.solasto.co.jp/ir/finance/highlight/'},
        ],
        'note': '業界は非上場プレイヤー多。ニチイ・ベネッセ・ツクイは非上場化済で開示が古い。上場でセグメント分離されているのはSOMPO・ソラスト・ベネッセ程度。',
    },

    # ===== 情報・通信 =====
    {
        'parent': '情報・通信',
        'segment': '通信キャリア (移動体)',
        'size_trillion': 19.0,
        'size_year': 'FY2024',
        'size_source': '総務省 通信市場・端末市場の動向 (令和7年6月) — MNO4社合計移動通信サービス売上 18兆9593億円',
        'size_url': 'https://www.soumu.go.jp/main_content/001014955.pdf',
        'companies': [
            {'name': 'KDDI (パーソナル)', 'revenue_oku': 50100, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 パーソナルセグメント (au通信+ライフデザイン領域含む)',
             'url': 'https://news.kddi.com/kddi/corporate/ir-news/2025/05/14/7636.html'},
            {'name': 'NTTドコモ', 'revenue_oku': 33545, 'fy': 'FY2024 (2025/3)', 'source': 'NTT 2024年度決算 ドコモ コンシューマ通信事業',
             'url': 'https://www.docomo.ne.jp/corporate/ir/binary/pdf/library/presentation/250509/presentation_fy2024_4q.pdf'},
            {'name': 'ソフトバンク (コンシューマ)', 'revenue_oku': 29529, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 コンシューマ事業 (モバイル+ブロードバンド)',
             'url': 'https://www.softbank.jp/corp/ir/documents/presentations/fy2024/q4_earnings_summary/'},
            {'name': '楽天モバイル (楽天G)', 'revenue_oku': 4407, 'fy': 'FY2024 (2024/12)', 'source': '2024年度 モバイルセグメント売上 (前年比+20.9%)',
             'url': 'https://corp.rakuten.co.jp/news/press/2025/0214_01.html'},
        ],
        'note': 'MNO4社のみ (5位該当なし)。KDDIパーソナルは金融・エネルギー等のライフデザイン領域を含むため純粋通信比較ではドコモ>SB>楽天が定説。',
    },
    {
        'parent': '情報・通信',
        'segment': 'SIer・ITサービス',
        'size_trillion': 7.02,
        'size_year': '2024年(暦年)',
        'size_source': 'IDC Japan 国内ITサービス市場予測2025 — 7兆205億円 (前年比+7.4%)',
        'size_url': 'https://my.idc.com/getdoc.jsp?containerId=prJPJ53253625',
        'companies': [
            {'name': 'NTTデータグループ', 'revenue_oku': 44300, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 連結売上収益 (国内+海外、海外比率高)',
             'url': 'https://www.nttdata.com/global/ja/-/media/nttdataglobal-ja/files/investors/library/earning/2024/4q/20254qbsr_all.pdf'},
            {'name': '日立製作所 (DSS)', 'revenue_oku': 28325, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 デジタルシステム&サービス',
             'url': 'https://www.hitachi.co.jp/New/cnews/month/2025/04/0428/2024_Anpre.pdf'},
            {'name': '富士通 (サービスソリューション)', 'revenue_oku': 22459, 'fy': 'FY2024 (2025/3)', 'source': '2024年度決算 サービスソリューション事業',
             'url': 'https://pr.fujitsu.com/jp/ir/library/presentation/pdf/20250424-02.pdf'},
            {'name': 'NEC (ITサービス)', 'revenue_oku': 20332, 'fy': 'FY2024 (2025/3)', 'source': '2024年度通期決算 ITサービスセグメント',
             'url': 'https://jpn.nec.com/ir/finance/segment.html'},
            {'name': '野村総合研究所', 'revenue_oku': 7648, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結売上収益',
             'url': 'https://www.nri.com/jp/ir/'},
        ],
        'note': 'NTTデータは海外含む連結値。次点はCTC(7282億)、TIS(5716億)、SCSK(約5000億)。',
    },

    # ===== 金融 =====
    {
        'parent': '金融',
        'segment': '損害保険 (元受正味保険料)',
        'size_trillion': 9.5,
        'size_year': 'FY2024',
        'size_source': '日本損害保険協会 ファクトブック2025 (2024年度元受正味保険料 約9.5兆円)',
        'size_url': 'https://www.sonpo.or.jp/report/publish/gyokai/ev7otb0000000061-att/fact2025_full.pdf',
        'companies': [
            {'name': '東京海上日動火災保険', 'revenue_oku': 25188, 'fy': 'FY2024 (2025/3)', 'source': '東京海上日動の現状2025 単体正味収入保険料',
             'url': 'https://www.tokiomarine-nichido.co.jp/company/pdf/TMNF_2025_d_05.pdf'},
            {'name': '損保ジャパン', 'revenue_oku': 22299, 'fy': 'FY2024 (2025/3)', 'source': '損保ジャパンの現状2025 単体正味収入保険料',
             'url': 'https://www.sompo-japan.co.jp/-/media/SJNK/files/company/disclosure/2025/sj_disc2025_07.pdf'},
            {'name': '三井住友海上火災保険', 'revenue_oku': 17000, 'fy': 'FY2024 (2025/3)', 'source': 'MS&AD 2024年度通期決算 単体 (FY2023実績+増収反映の概算)',
             'url': 'https://www.ms-ad-hd.com/ja/news/irnews/auto_20250508535356/pdfFile.pdf'},
            {'name': 'あいおいニッセイ同和損保', 'revenue_oku': 14500, 'fy': 'FY2024 (2025/3)', 'source': 'MS&AD 2024年度通期決算 単体 (FY2023実績+増収反映の概算)',
             'url': 'https://www.ms-ad-hd.com/ja/news/irnews/auto_20250508535356/pdfFile.pdf'},
        ],
        'note': '3メガで業界の8割超を寡占。MS&AD国内損保2社+三井ダイレクト合算で約3.2兆円。',
    },
    {
        'parent': '金融',
        'segment': '生命保険 (保険料等収入)',
        'size_trillion': 37.52,
        'size_year': 'FY2024',
        'size_source': '生命保険協会 2025年版 生命保険の動向 — 収入保険料37兆5217億円 (前年度比108.8%)',
        'size_url': 'https://www.seiho.or.jp/data/statistics/trend/pdf/all_2025.pdf',
        'companies': [
            {'name': 'かんぽ生命保険', 'revenue_oku': 67441, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 連結保険料等収入 (日本郵政G)',
             'url': 'https://www.jp-life.japanpost.jp/information/press/2025/abt_prs_id002060.html'},
            {'name': '日本生命保険', 'revenue_oku': 47946, 'fy': 'FY2024 (2025/3)', 'source': '2024年度業績概要 単体保険料等収入',
             'url': 'https://www.nissay.co.jp/kaisha/annai/gyoseki/pdf/kessan202505/gaiyo.pdf'},
            {'name': '明治安田生命保険', 'revenue_oku': 27583, 'fy': 'FY2024 (2025/3)', 'source': '2024年度決算 単体保険料等収入',
             'url': 'https://www.meijiyasuda.co.jp/profile/corporate_info/disclosure/account/2024/pdf/close_2025_point.pdf'},
            {'name': '第一フロンティア生命', 'revenue_oku': 22596, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 保険料収入 (第一生命HD傘下)',
             'url': 'https://www.dai-ichi-life.co.jp/company/results/kessan/2024/pdf/index_001.pdf'},
            {'name': '第一生命保険', 'revenue_oku': 19500, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 単体 (第一生命HD連結6.80兆-子会社控除の概数)',
             'url': 'https://www.dai-ichi-life.co.jp/company/results/kessan/2024/pdf/index_001.pdf'},
        ],
        'note': '住友生命G保険料等収入3.38兆円 (連結) は次点級。第一生命単体FY2024は概数推定。',
    },

    # ===== エネルギー =====
    {
        'parent': 'エネルギー',
        'segment': '電力小売',
        'size_trillion': 18.0,
        'size_year': 'FY2024',
        'size_source': '資源エネルギー庁 電力調査統計 (販売電力量821.8TWh×加重平均22円/kWh で約18兆円推計)',
        'size_url': 'https://www.enecho.meti.go.jp/statistics/electric_power/ep002/',
        'companies': [
            {'name': '東京電力HD (エナジーパートナー)', 'revenue_oku': 53700, 'fy': 'FY2024 (2025/3)', 'source': '2024年度決算 エナジーパートナー(EP) セグメント',
             'url': 'https://www.tepco.co.jp/about/ir/library/results/pdf/2503q4gaiyou-j.pdf'},
            {'name': '関西電力 (エネルギー事業)', 'revenue_oku': 35400, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 エネルギー事業セグメント (発電含む)',
             'url': 'https://www.kepco.co.jp/corporate/pr/2025/pdf/20250430_6j.pdf'},
            {'name': '中部電力 (ミライズ)', 'revenue_oku': 29100, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 ミライズ (小売)',
             'url': 'https://www.chuden.co.jp/ir/ir_siryo/kessan/__icsFiles/afieldfile/2025/04/28/2024tanshin_4qua.pdf'},
            {'name': '東北電力', 'revenue_oku': 26449, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結売上 (小売主体)',
             'url': 'https://www.tohoku-epco.co.jp/ir_n/report/finance_results/pdf/2025_tan.pdf'},
            {'name': '九州電力', 'revenue_oku': 23568, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結売上',
             'url': 'https://www.kyuden.co.jp/var/rev0/0749/6297/Kf4JR2pt.pdf'},
        ],
        'note': '市場規模は数量×推計単価。各社のセグメント定義 (小売/エネルギー事業/連結) が不揃いで直接比較は困難。',
    },
    {
        'parent': 'エネルギー',
        'segment': '都市ガス小売',
        'size_trillion': 4.5,
        'size_year': 'FY2024',
        'size_source': '資源エネルギー庁 ガス事業生産動態統計 + 業界推計 (都市ガス小売販売金額4-5兆円)',
        'size_url': 'https://www.enecho.meti.go.jp/statistics/gas/ga001/results.html',
        'companies': [
            {'name': '東京ガス', 'revenue_oku': 23100, 'fy': 'FY2024 (2025/3)', 'source': '2024年度決算 エネルギー・ソリューション (都市ガス+電力+LNG)',
             'url': 'https://www.tokyo-gas.co.jp/news/press/20250428-01.pdf'},
            {'name': '大阪ガス', 'revenue_oku': 17379, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 国内エネルギー (都市ガス+電力)',
             'url': 'https://www.daigasgroup.com/ir/library/earnings/__icsFiles/afieldfile/2025/05/07/20250508_4.pdf'},
            {'name': '東邦ガス', 'revenue_oku': 4293, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 ガス事業セグメント',
             'url': 'https://www.tohogas.co.jp/corporate-n/press/__icsFiles/afieldfile/2025/04/30/press20250430_1.pdf'},
            {'name': '西部ガスHD', 'revenue_oku': 1700, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上 (概数)',
             'url': 'https://www.saibugas.co.jp/company/ir/'},
            {'name': '北海道ガス', 'revenue_oku': 1100, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上 (概数)',
             'url': 'https://www.hokkaido-gas.co.jp/ir/'},
        ],
        'note': '東京ガス・大阪ガスのセグメントは電力・LNGトレーディング含む広い定義 (純都市ガス小売より広い)。',
    },

    # ===== 運輸 =====
    {
        'parent': '運輸',
        'segment': '宅配・路線便',
        'size_trillion': 2.5,
        'size_year': 'FY2024',
        'size_source': '国交省 令和6年度宅配便取扱実績 50.31億個。市場金額は主要4社合算で2.5兆円規模 (路線便含むと約3.5兆円)',
        'size_url': 'https://www.mlit.go.jp/report/press/jidosha04_hh_000341.html',
        'companies': [
            {'name': '日本郵政 (郵便・物流)', 'revenue_oku': 20808, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 郵便・物流事業セグメント (郵便+ゆうパック+国際物流)',
             'url': 'https://www.japanpost.jp/ir/library/earnings/pdf/20250515_03.pdf'},
            {'name': 'ヤマトHD', 'revenue_oku': 15347, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 エクスプレス事業セグメント',
             'url': 'https://www.yamato-hd.co.jp/investors/library/results/'},
            {'name': 'SGホールディングス', 'revenue_oku': 10211, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 デリバリー事業セグメント (飛脚宅配便)',
             'url': 'https://www.sg-hldgs.co.jp/ir/library/'},
            {'name': 'セイノーHD', 'revenue_oku': 5541, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 輸送事業セグメント (特積み・路線便)',
             'url': 'https://www.seino.co.jp/seino/media/pdf-lib/20250514-05.pdf'},
            {'name': '福山通運', 'revenue_oku': 2345, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 運送事業セグメント (特積み路線便専業)',
             'url': 'https://www.fukutsu.co.jp/ir/'},
        ],
        'note': 'ヤマト/SGは宅配便特化、セイノー/福山は特積み (路線便) 特化。日本郵便は単体未開示で日本郵政G郵便・物流セグメントを採用。',
    },

    # ===== 製造業 =====
    {
        'parent': '製造業',
        'segment': '半導体製造装置 (日本市場)',
        'size_trillion': 1.22,
        'size_year': 'FY2024',
        'size_source': 'SEAJ 2025年1月需要予測 — 日本市場販売 1兆2232億円 (前年比+7%)',
        'size_url': 'https://www.seaj.or.jp/file/seajforecastjan2025_japanese_for%20press.pdf',
        'companies': [
            {'name': '東京エレクトロン', 'revenue_oku': 24316, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結売上高 (グローバル)',
             'url': 'https://www.tel.co.jp/ir/library/report/index.html'},
            {'name': 'アドバンテスト', 'revenue_oku': 5981, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 半導体・部品テストシステム (グローバル)',
             'url': 'https://www.advantest.com/ja/investors/ir-library/result/'},
            {'name': 'SCREEN HD', 'revenue_oku': 5195, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期 SPE (半導体製造装置) セグメント',
             'url': 'https://www.screen.co.jp/ir/'},
            {'name': 'ディスコ', 'revenue_oku': 3933, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結売上 (グローバル)',
             'url': 'https://www.disco.co.jp/jp/ir/library/doc/fr/fr20250417.pdf'},
            {'name': 'KOKUSAI ELECTRIC', 'revenue_oku': 2389, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 連結売上収益 (グローバル)',
             'url': 'https://www.kokusai-electric.com/ir/library/result'},
        ],
        'note': '【重要】日本市場規模1.22兆円 vs 各社売上はグローバル連結 (輸出比率7-8割)。日本のみ売上は各社不開示。日系メーカーの世界販売合計は約4.4兆円。',
    },
    {
        'parent': '製造業',
        'segment': '高炉鉄鋼',
        'size_trillion': 23.8,
        'size_year': '2024年(暦年)',
        'size_source': '経構造実態調査 (製造業) 2024年 — 鉄鋼業 製造品出荷額 23兆8346億円',
        'size_url': 'https://www.meti.go.jp/statistics/tyo/kkj/pdf/seizo_gaikyo2024.pdf',
        'companies': [
            {'name': '日本製鉄', 'revenue_oku': 78743, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信(IFRS) 製鉄事業セグメント',
             'url': 'https://www.nipponsteel.com/ir/library/settlement/pdf/20250509_200.pdf'},
            {'name': 'JFEホールディングス', 'revenue_oku': 33651, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信(IFRS) 鉄鋼事業セグメント',
             'url': 'https://www.jfe-holdings.co.jp/uploads/2024-all250508.pdf'},
            {'name': '神戸製鋼所', 'revenue_oku': 11161, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信 鉄鋼アルミ事業セグメント (鋼材内数9144億)',
             'url': 'https://www.kobelco.co.jp/ir/results/pdf/250512_kessan.pdf'},
            {'name': '大同特殊鋼', 'revenue_oku': 5800, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算短信(IFRS) 連結売上収益',
             'url': 'https://www.daido.co.jp/common/pdf/pages/ir/library/result/2503_4q.pdf'},
            {'name': '愛知製鋼', 'revenue_oku': 2993, 'fy': 'FY2024 (2025/3)', 'source': '2025年3月期決算 連結売上収益',
             'url': 'https://www.aichi-steel.co.jp/ir/'},
        ],
        'note': '市場規模は製造品出荷額 (粗鋼+下流加工含む)。粗鋼生産量は2024年度8295万トン。',
    },
]


# ===== Workbook build =====
def build():
    wb = Workbook()
    ws = wb.active
    ws.title = '1兆円市場マップ'

    # Title row
    ws.merge_cells('A1:P1')
    ws['A1'] = '日本の1兆円市場マップ (FY2024) — 親産業 / サブセグメント / 市場規模 / 上位5社のセグメント売上'
    ws['A1'].font = TITLE
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26

    # Column widths
    widths = {
        'A': 12, 'B': 26, 'C': 12,
        'D': 22, 'E': 13,
        'F': 22, 'G': 13,
        'H': 22, 'I': 13,
        'J': 22, 'K': 13,
        'L': 22, 'M': 13,
        'N': 38, 'O': 38, 'P': 32,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header row (row 3)
    headers = [
        '親産業', 'サブセグメント', '市場規模(兆円)',
        '1位 社名', '1位 売上(億円)',
        '2位 社名', '2位 売上(億円)',
        '3位 社名', '3位 売上(億円)',
        '4位 社名', '4位 売上(億円)',
        '5位 社名', '5位 売上(億円)',
        '出典 (市場規模)', '出典 (企業売上代表例)', '注記',
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 36

    # Data rows
    row = 4
    prev_parent = None
    for m in MARKETS:
        parent_changed = (m['parent'] != prev_parent)
        prev_parent = m['parent']

        cells = []
        cells.append(m['parent'])
        cells.append(m['segment'])
        cells.append(m.get('size_trillion'))

        comps = m.get('companies', [])
        for i in range(5):
            if i < len(comps):
                cells.append(comps[i]['name'])
                cells.append(comps[i].get('revenue_oku'))
            else:
                cells.append(None)
                cells.append(None)

        cells.append(m.get('size_source', ''))
        cells.append(_company_sources_summary(comps))
        cells.append(m.get('note', ''))

        for col_idx, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = BLACK
            c.alignment = Alignment(
                horizontal='left' if col_idx not in (3, 5, 7, 9, 11, 13) else 'right',
                vertical='center',
                wrap_text=(col_idx >= 14)
            )
            c.border = BORDER
            if parent_changed and col_idx == 1:
                c.fill = PARENT_FILL
                c.font = BLACK_B
            elif col_idx in (1, 2):
                c.fill = ALT_FILL

        # Number formats
        ws.cell(row=row, column=3).number_format = FMT_TRN
        for col_idx in (5, 7, 9, 11, 13):
            ws.cell(row=row, column=col_idx).number_format = FMT_OKU

        # Hyperlink for size source URL
        url = m.get('size_url')
        if url:
            ws.cell(row=row, column=14).hyperlink = url
            ws.cell(row=row, column=14).font = Font(color='0563C1', name='Yu Gothic', size=10, underline='single')

        ws.row_dimensions[row].height = 42
        row += 1

    # Freeze panes (header + first 2 cols)
    ws.freeze_panes = 'C4'

    # Auto filter
    ws.auto_filter.ref = f'A3:P{row-1}'

    out_dir = Path(__file__).parent / 'fermi-estimation' / 'output'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / '日本の1兆円市場マップ_FY2024.xlsx'
    wb.save(out_path)
    print(f'Saved: {out_path}')
    print(f'Markets: {len(MARKETS)} rows')
    return out_path


def _company_sources_summary(comps):
    """企業売上の出典URLを改行区切りで結合 (1社目から最大5社)"""
    lines = []
    for c in comps[:5]:
        url = c.get('url', '')
        src = c.get('source', '')
        nm = c.get('name', '')
        line = f"{nm}: {src}" + (f" / {url}" if url else "")
        lines.append(line)
    return "\n".join(lines)


if __name__ == '__main__':
    build()
