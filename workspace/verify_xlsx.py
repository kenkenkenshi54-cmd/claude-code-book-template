# -*- coding: utf-8 -*-
"""Verify the generated xlsx by reading key cells."""
from openpyxl import load_workbook

path = r'c:\Users\Kamei.Kenshi\Documents\dev\claude-code-book-template\workspace\ビルメンテナンス業界_市場規模推定.xlsx'
wb = load_workbook(path)
ws = wb.active
print('Sheet:', ws.title)
print()
print('=== Row 7 (Adopted) ===')
for col in ['B','G','H','I','J','K','L','M','N','O','P','Q','S','T']:
    v = ws[f'{col}7'].value
    print(f'{col}7 = {v}')

print()
print('=== Row 10-13 (Triangulation) ===')
for r in [10,11,12,13]:
    print(f'B{r}={ws[f"B{r}"].value}  K{r}={ws[f"K{r}"].value}  U{r}={ws[f"U{r}"].value}')

print()
print('=== Row 22-24 (Estimate 2 detail) ===')
for r in [22,23,24]:
    print(f'B{r}={ws[f"B{r}"].value}  K{r}={ws[f"K{r}"].value}')

print()
print('=== Row 26-32 (Estimate 3 detail) ===')
for r in [26,27,28,29,30,31,32]:
    print(f'B{r}={ws[f"B{r}"].value}  K{r}={ws[f"K{r}"].value}')

print()
print('=== Row 36-47 (CAGR analysis) ===')
for r in [36,37,38,40,41,42,43,44,45,46,47]:
    print(f'B{r}={ws[f"B{r}"].value}  K{r}={ws[f"K{r}"].value}')

# Now compute manually for sanity
print()
print('=== Sanity Computations ===')
# K22 = K23/K24 = 856/0.018
print('K22 (=K23/K24):', 856/0.018)
# K26 = K31/K32 = (3379+856+751+1387)/0.135
total = 3379+856+751+1387
print('K31 sum:', total)
print('K26 (=K31/K32):', total/0.135)
# Past CAGR = (51615/43425)^(1/3)-1
past_cagr = (51615/43425)**(1/3)-1
print(f'Past CAGR (S7): {past_cagr*100:.2f}%')
# Future CAGR L=52685, Q=52685*1.02^5
Q = 52685 * (1.02**5)
print(f'Q7 computed: {Q:.0f}')
fut_cagr = (Q/52685)**(1/5)-1
print(f'Future CAGR (T7): {fut_cagr*100:.2f}%')
# Divergence
print(f'Divergence: {(fut_cagr - past_cagr)*100:.2f}pp')
# Factor sum
factors = 0.3+0.2-2.0-0.5-1.0
print(f'Factor sum: {factors:+.2f}pp')
print(f'Residual: {(fut_cagr - past_cagr)*100 - factors:+.2f}pp')

# Save status marker
with open(r'c:\Users\Kamei.Kenshi\Documents\dev\claude-code-book-template\workspace\verify_done.txt', 'w', encoding='utf-8') as f:
    f.write('OK')
print('VERIFY DONE')
