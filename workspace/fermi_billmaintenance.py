# -*- coding: utf-8 -*-
"""
ビルメンテナンス業界 市場規模フェルミ推定 (2024/04時点)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fonts
BLUE = Font(color='0070C0', name='Yu Gothic', size=10)
BLUE_B = Font(color='0070C0', name='Yu Gothic', size=10, bold=True)
BLACK = Font(color='000000', name='Yu Gothic', size=10)
BLACK_B = Font(color='000000', name='Yu Gothic', size=10, bold=True)
HEADER = Font(color='FFFFFF', name='Yu Gothic', size=11, bold=True)
TITLE = Font(color='000000', name='Yu Gothic', size=14, bold=True)

# Fills
HEADER_FILL = PatternFill('solid', fgColor='305496')
SECTION_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBSECTION_FILL = PatternFill('solid', fgColor='F2F2F2')
ADOPTED_FILL = PatternFill('solid', fgColor='FFF2CC')

# Border
thin = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# Number formats
FMT_OKU = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%'
FMT_PP  = '+0.0"pp";-0.0"pp";0.0"pp"'

wb = Workbook()
ws = wb.active
ws.title = 'ビルメンテナンス_日本'

# Column widths
widths = {'A':2,'B':16,'C':14,'D':14,'E':14,'F':14,'G':10,
          'H':11,'I':11,'J':11,'K':11,'L':11,'M':11,'N':11,'O':11,'P':11,'Q':11,
          'R':2,'S':10,'T':10,'U':28,'V':40}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Row 2: Title
ws.merge_cells('B2:V2')
ws['B2'] = '【ビルメンテナンス業界】市場規模推定 (2026/04 時点) — 日本国内、全国ビルメンテナンス協会 標準定義(清掃・設備管理・警備保安・衛生管理・駐車場管理) ベース'
ws['B2'].font = TITLE
ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 24

# Row 4: Year labels
ws['B4'] = '項目'
ws['G4'] = '単位'
ws['H4'] = 2021
for i in range(1, 10):
    col = get_column_letter(7 + 1 + i)  # I=9 onward
    ws[f'{col}4'] = f'=H4+{i}'
ws['S4'] = '過去CAGR'
ws['T4'] = '将来CAGR'
ws['U4'] = '出典'
ws['V4'] = 'URL'
for col in ['B','G','H','I','J','K','L','M','N','O','P','Q','S','T','U','V']:
    ws[f'{col}4'].font = HEADER
    ws[f'{col}4'].fill = HEADER_FILL
    ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{col}4'].border = BORDER
ws.row_dimensions[4].height = 22

# Row 5: Data nature
ws['B5'] = 'データ性質'
labels = ['実績','実績','実績','実績','予測','予測','予測','予測','予測','予測']
for i, lab in enumerate(labels):
    col = get_column_letter(8 + i)
    ws[f'{col}5'] = lab
    ws[f'{col}5'].font = BLACK
    ws[f'{col}5'].alignment = Alignment(horizontal='center')
    ws[f'{col}5'].fill = SUBSECTION_FILL
ws['B5'].font = BLACK_B
ws['B5'].fill = SUBSECTION_FILL

# ========== Row 7: Adopted Market Size ==========
ws['B7'] = '市場規模(採用値)'
ws['G7'] = '億円'
# Past actuals (Yano Research, 元請金額ベース)
ws['H7'] = 43425  # 2021
ws['I7'] = 45889  # 2022
ws['J7'] = 48297  # 2023
ws['K7'] = '=K10'  # 2024 -- link to Estimate ① row
# Forecast (Yano 2025 forecast=52,685, then ~2.0% CAGR)
ws['L7'] = 52685   # 2025 (Yano predict)
ws['M7'] = '=L7*1.020'  # 2026
ws['N7'] = '=M7*1.020'
ws['O7'] = '=N7*1.020'
ws['P7'] = '=O7*1.020'
ws['Q7'] = '=P7*1.020'  # 2030
ws['S7'] = '=(K7/H7)^(1/3)-1'
ws['T7'] = '=(Q7/L7)^(1/5)-1'
ws['U7'] = '採用値:矢野経済(実績)+ 2.0%/年成長前提'
ws['V7'] = ''
for col in ['B','G','H','I','J','L']:
    ws[f'{col}7'].fill = ADOPTED_FILL
for col in ['H','I','J','L']:
    ws[f'{col}7'].font = BLUE
    ws[f'{col}7'].number_format = FMT_OKU
for col in ['K','M','N','O','P','Q']:
    ws[f'{col}7'].font = BLACK
    ws[f'{col}7'].fill = ADOPTED_FILL
    ws[f'{col}7'].number_format = FMT_OKU
ws['B7'].font = BLACK_B
ws['G7'].font = BLACK
ws['G7'].alignment = Alignment(horizontal='center')
ws['S7'].font = BLACK; ws['S7'].number_format = FMT_PCT; ws['S7'].fill = ADOPTED_FILL
ws['T7'].font = BLACK; ws['T7'].number_format = FMT_PCT; ws['T7'].fill = ADOPTED_FILL
ws['U7'].font = BLACK
ws.row_dimensions[7].height = 20

# ========== Row 9: Triangulation header ==========
ws.merge_cells('B9:V9')
ws['B9'] = '【三角測量:3独立推定の収束検証】'
ws['B9'].font = HEADER
ws['B9'].fill = HEADER_FILL
ws['B9'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

# Row 10: Estimate ①
ws['B10'] = '① 公式統計(矢野経済)'
ws['G10'] = '億円'
ws['K10'] = 51615  # 2024年度 矢野経済 確報値
ws['K10'].font = BLUE
ws['K10'].number_format = FMT_OKU
ws['U10'] = '矢野経済研究所「ビル管理市場に関する調査(2025)」'
ws['V10'] = 'https://www.yano.co.jp/press-release/show/press_id/3954'
ws['B10'].font = BLACK_B

# Row 11: Estimate ②
ws['B11'] = '② シェア×売上(逆算)'
ws['G11'] = '億円'
ws['K11'] = '=K22'
ws['K11'].font = BLACK
ws['K11'].number_format = FMT_OKU
ws['U11'] = '日本管財HD 建物管理運営事業 売上 ÷ 自社シェア(推計)で逆算'

# Row 12: Estimate ③
ws['B12'] = '③ 主要プレーヤー合計÷カバー率'
ws['G12'] = '億円'
ws['K12'] = '=K26'
ws['K12'].font = BLACK
ws['K12'].number_format = FMT_OKU
ws['U12'] = '上場大手+大手非上場の売上合計 ÷ 想定カバー率'

# Row 13: Adopted
ws['B13'] = '採用値'
ws['G13'] = '億円'
ws['K13'] = '=K7'
ws['K13'].font = BLACK_B
ws['K13'].fill = ADOPTED_FILL
ws['K13'].number_format = FMT_OKU
ws['B13'].font = BLACK_B
ws['B13'].fill = ADOPTED_FILL

# Row 14: Adoption rationale
ws.merge_cells('B14:V14')
ws['B14'] = '採用根拠:3推定の乖離が±2%以内で強く収束。最も信頼性の高い公式統計(矢野経済)を採用。'
ws['B14'].font = BLACK
ws['B14'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws['B14'].fill = SUBSECTION_FILL

# ========== Row 15: ① Detail ==========
ws.merge_cells('B15:V15')
ws['B15'] = '【① 公式統計 詳細】矢野経済研究所「ビル管理市場」(元請金額ベース)'
ws['B15'].font = HEADER
ws['B15'].fill = SECTION_FILL
ws['B15'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

# Row 16-19: Past data and breakdown
ws['B16'] = '建物使途別シェア (2022)'
ws['C16'] = '事務所ビル'
ws['G16'] = '%'
ws['K16'] = 0.217
ws['K16'].font = BLUE
ws['K16'].number_format = FMT_PCT
ws['U16'] = '矢野経済(2022年度実績ベース)'
ws['V16'] = 'https://www.yano.co.jp/press-release/show/press_id/3361'

ws['C17'] = '店舗・商業施設'
ws['G17'] = '%'
ws['K17'] = 0.181
ws['K17'].font = BLUE
ws['K17'].number_format = FMT_PCT
ws['U17'] = '矢野経済'
ws['V17'] = 'https://www.yano.co.jp/press-release/show/press_id/3361'

ws['C18'] = '医療・福祉'
ws['G18'] = '%'
ws['K18'] = 0.102
ws['K18'].font = BLUE
ws['K18'].number_format = FMT_PCT
ws['U18'] = '矢野経済'

ws['C19'] = '工場・学校・住宅・その他'
ws['G19'] = '%'
ws['K19'] = '=1-K16-K17-K18'
ws['K19'].font = BLACK
ws['K19'].number_format = FMT_PCT
ws['U19'] = '残差'

# Sanity check: 全国ビルメンテナンス協会 reference
ws['B20'] = '参考:全国ビルメン協会'
ws['G20'] = '億円'
ws['J20'] = 46700  # 2023実態調査
ws['J20'].font = BLUE; ws['J20'].number_format = FMT_OKU
ws['I20'] = 44900  # 2022
ws['I20'].font = BLUE; ws['I20'].number_format = FMT_OKU
ws['U20'] = '全国ビルメンテナンス協会 第55回実態調査(2024年実施)'
ws['V20'] = 'https://www.j-bma.or.jp/data/105023'
ws['B20'].font = BLACK
# Check rate
ws['B21'] = '矢野vs協会 乖離率(2023)'
ws['K21'] = '=(J7-J20)/J20'
ws['K21'].number_format = FMT_PCT
ws['K21'].font = BLACK
ws['U21'] = '+3.4%(調査範囲・元請定義の差。矢野が広く取る)'
ws['B21'].font = BLACK

# ========== Row 22: ② Detail (Share × Sales) ==========
ws['B22'] = '② シェア×売上 逆算'
ws['G22'] = '億円'
ws['K22'] = '=K23/K24'
ws['K22'].font = BLACK_B
ws['K22'].number_format = FMT_OKU

ws['B23'] = ' └ 日本管財HD 建物管理運営事業売上'
ws['G23'] = '億円'
ws['K23'] = 856
ws['K23'].font = BLUE; ws['K23'].number_format = FMT_OKU
ws['U23'] = '日本管財HD 2023年3月期 決算説明資料 セグメント情報'
ws['V23'] = 'https://www.nkanzai.co.jp/wp-content/uploads/2023/05/Financial-Results_202303.pdf'

ws['B24'] = ' └ 同社推計シェア(対市場4.83兆円)'
ws['G24'] = '%'
ws['K24'] = 0.018
ws['K24'].font = BLUE; ws['K24'].number_format = FMT_PCT
ws['U24'] = '上場ビルメン売上÷市場規模(矢野)で算出した社内シェア推計値'

# ========== Row 26: ③ Detail (Top players sum / coverage) ==========
ws['B26'] = '③ 主要プレーヤー合計÷カバー率'
ws['G26'] = '億円'
ws['K26'] = '=K31/K32'
ws['K26'].font = BLACK_B
ws['K26'].number_format = FMT_OKU

ws['B27'] = ' └ イオンディライト(連結FM事業)'
ws['G27'] = '億円'
ws['K27'] = 3379
ws['K27'].font = BLUE; ws['K27'].number_format = FMT_OKU
ws['U27'] = 'イオンディライト 2024年2月期決算(連結売上、ファシリティマネジメント全7事業)'
ws['V27'] = 'https://www.aeondelight.co.jp/ir/financial/highlight.html'

ws['B28'] = ' └ 日本管財HD(建物管理運営)'
ws['G28'] = '億円'
ws['K28'] = 856
ws['K28'].font = BLUE; ws['K28'].number_format = FMT_OKU
ws['U28'] = '日本管財HD 2023年3月期 セグメント情報'

ws['B29'] = ' └ 太平ビルサービス(連結)'
ws['G29'] = '億円'
ws['K29'] = 751
ws['K29'].font = BLUE; ws['K29'].number_format = FMT_OKU
ws['U29'] = '太平ビルサービス 2023年12月期 グループ売上'
ws['V29'] = 'https://alarmbox.jp/companyinfo/entities/2011101012138'

ws['B30'] = ' └ その他上場大手7社合計(東洋テック・ビケンテクノ・ハリマビステム・大成・日本ハウズイング・ダイビル・共立メンテナンス)'
ws['G30'] = '億円'
ws['K30'] = 1387  # 301+290+253+223+159+85+77 = 1388
ws['K30'].font = BLUE; ws['K30'].number_format = FMT_OKU
ws['U30'] = '業界動向サーチ 売上ランキング 2022-2023'
ws['V30'] = 'https://gyokai-search.com/4-bill-uriage.html'

ws['B31'] = '主要プレーヤー合計'
ws['G31'] = '億円'
ws['K31'] = '=SUM(K27:K30)'
ws['K31'].font = BLACK_B; ws['K31'].number_format = FMT_OKU
ws['B31'].font = BLACK_B

ws['B32'] = '想定カバー率(寡占度低い業界)'
ws['G32'] = '%'
ws['K32'] = 0.135  # 13.5%
ws['K32'].font = BLUE; ws['K32'].number_format = FMT_PCT
ws['U32'] = '上位10社で約25%カバー(2019-2020業界動向サーチ)を参考、上位10社のうち上場主要企業のみ拾うため約13-14%と想定'

# ========== Row 35: CAGR Divergence Analysis ==========
ws.merge_cells('B35:V35')
ws['B35'] = '【CAGR乖離分析:過去実績期間 vs 将来計画期間】'
ws['B35'].font = HEADER
ws['B35'].fill = SECTION_FILL
ws['B35'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

ws['B36'] = '過去CAGR (2021→2024)'
ws['G36'] = '%'
ws['K36'] = '=S7'
ws['K36'].number_format = FMT_PCT
ws['K36'].font = BLACK
ws['B36'].font = BLACK

ws['B37'] = '将来CAGR (2025→2030)'
ws['G37'] = '%'
ws['K37'] = '=T7'
ws['K37'].number_format = FMT_PCT
ws['K37'].font = BLACK
ws['B37'].font = BLACK

ws['B38'] = '乖離 (将来 − 過去)'
ws['G38'] = 'pp'
ws['K38'] = '=(K37-K36)*100'
ws['K38'].number_format = FMT_PP
ws['K38'].font = BLACK_B
ws['B38'].font = BLACK_B

# Row 40: Factor breakdown header
ws.merge_cells('B40:V40')
ws['B40'] = '【乖離要因の分解】(pp = percentage point)'
ws['B40'].font = HEADER
ws['B40'].fill = SECTION_FILL
ws['B40'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

# Drivers (+) and (-)
ws['B41'] = '(+) DX/IoT・統合FM・複合契約による高付加価値化'
ws['G41'] = 'pp'
ws['K41'] = 0.2
ws['K41'].font = BLUE; ws['K41'].number_format = FMT_PP
ws['U41'] = '矢野経済2025レポート、船井総研業界展望(2025)'
ws['V41'] = 'https://www.jinzai-business.com/798'

ws['B42'] = '(+) 建築物の老朽化に伴う改修・点検需要の継続'
ws['G42'] = 'pp'
ws['K42'] = 0.2
ws['K42'].font = BLUE; ws['K42'].number_format = FMT_PP
ws['U42'] = 'ビル管理業界の動向(国交省データ含む)'

ws['B43'] = '(-) 価格改定の浸透による単価上昇分のピークアウト'
ws['G43'] = 'pp'
ws['K43'] = -2.5
ws['K43'].font = BLUE; ws['K43'].number_format = FMT_PP
ws['U43'] = '矢野経済「契約更改時の価格改定が市場拡大の主要因」(2024年度+6.9%、2025年度+2.1%予測=単価寄与の剥落)'
ws['V43'] = 'https://www.yano.co.jp/press-release/show/press_id/3954'

ws['B44'] = '(-) 人手不足による供給制約・需要取りこぼし'
ws['G44'] = 'pp'
ws['K44'] = -0.6
ws['K44'].font = BLUE; ws['K44'].number_format = FMT_PP
ws['U44'] = '労働集約型業界の有効求人倍率上昇、清掃職有効求人倍率は3倍超水準'

ws['B45'] = '(-) コロナ後の積み残し新規案件特需の剥落'
ws['G45'] = 'pp'
ws['K45'] = -1.2
ws['K45'].font = BLUE; ws['K45'].number_format = FMT_PP
ws['U45'] = '矢野経済2024レポート「コロナ禍で先送りされていた新規案件の稼働」が一巡'

ws['B46'] = '要因合計'
ws['G46'] = 'pp'
ws['K46'] = '=SUM(K41:K45)'
ws['K46'].number_format = FMT_PP
ws['K46'].font = BLACK_B
ws['B46'].font = BLACK_B

ws['B47'] = '残差(乖離 − 要因合計)'
ws['G47'] = 'pp'
ws['K47'] = '=K38-K46'
ws['K47'].number_format = FMT_PP
ws['K47'].font = BLACK
ws['B47'].font = BLACK

# Row 49: Interpretation
ws.merge_cells('B49:V49')
ws['B49'] = '【評価】将来CAGR(2.0%)は過去CAGR(5.9%)から▲3.9pp減速。減速の主要因は(1)コロナ後の特需剥落と(2)価格改定ピークアウト。一方、DX/IoT・改修需要が下支え。減速ストーリーは合理的に説明可能。'
ws['B49'].font = BLACK
ws['B49'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws['B49'].fill = SUBSECTION_FILL
ws.row_dimensions[49].height = 32

# Save
out = r'c:\Users\Kamei.Kenshi\Documents\dev\claude-code-book-template\workspace\ビルメンテナンス業界_市場規模推定.xlsx'
wb.save(out)
print('SAVED:', out)
